from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.contrib.auth import get_user_model
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta

from .forms import ConsultaForm
from .models import Consulta, NotificacionWhatsApp
from pacientes.models import Paciente

import openpyxl
from xhtml2pdf import pisa


# ----------------------------
# Vista para registrar una nueva consulta
# ----------------------------
@login_required
def registrar_consulta(request):
    if request.method == 'POST':
        form = ConsultaForm(request.POST)
        if form.is_valid():
            consulta = form.save(commit=False)
            consulta.doctor = request.user
            
            # Calcular duración automáticamente si no se especifica
            if not consulta.duracion_consulta:
                consulta.duracion_consulta = consulta.calcular_duracion_consulta()
            
            consulta.save()
            messages.success(request, f"✅ Consulta registrada exitosamente para {consulta.paciente.nombre_completo()}.")
            return redirect('listar_consultas')
        else:
            messages.error(request, "❌ Hubo un error al registrar la consulta. Por favor revisa los datos.")
    else:
        form = ConsultaForm()
    
    # Obtener estadísticas para mostrar en el formulario
    total_pacientes = Paciente.objects.filter(estado='Activo').count()
    consultas_hoy = Consulta.objects.filter(fecha__date=timezone.now().date()).count()
    
    context = {
        'form': form,
        'total_pacientes': total_pacientes,
        'consultas_hoy': consultas_hoy,
    }
    return render(request, 'consultas/registrar_consulta.html', context)


# ----------------------------
# Vista del Dashboard mejorado
# ----------------------------
@login_required
def dashboard(request):
    # Estadísticas básicas
    total_consultas = Consulta.objects.count()
    total_pacientes = Paciente.objects.filter(estado='Activo').count()
    usuario = request.user.get_full_name() or request.user.username
    
    # Estadísticas por fecha
    hoy = timezone.now().date()
    consultas_hoy = Consulta.objects.filter(fecha__date=hoy).count()
    consultas_semana = Consulta.objects.filter(fecha__date__gte=hoy - timedelta(days=7)).count()
    consultas_mes = Consulta.objects.filter(fecha__date__gte=hoy - timedelta(days=30)).count()
    
    # Consultas de emergencia
    consultas_emergencia = Consulta.objects.filter(tipo_consulta='Emergencia').count()
    
    # Duración promedio
    duracion_promedio = Consulta.objects.aggregate(
        promedio=Avg('duracion_consulta')
    )['promedio'] or 30
    
    # Consultas recientes
    ultimas_consultas = Consulta.objects.select_related('paciente').order_by('-fecha')[:5]
    
    # Estadísticas por tipo de consulta
    tipos_consulta = Consulta.objects.values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Pacientes con más consultas
    pacientes_frecuentes = Paciente.objects.annotate(
        total_consultas=Count('consultas')
    ).filter(total_consultas__gt=0).order_by('-total_consultas')[:5]
    
    # Alertas médicas recientes
    consultas_recientes = Consulta.objects.select_related('paciente').filter(
        fecha__date__gte=hoy - timedelta(days=7)
    )
    alertas = []
    for consulta in consultas_recientes:
        alertas.extend(consulta.obtener_alertas_medicas())
    
    # Próximas citas
    proximas_citas = Consulta.objects.select_related('paciente').filter(
        proxima_cita__gte=timezone.now()
    ).order_by('proxima_cita')[:5]

    # Fecha actual formateada
    from datetime import datetime
    fecha_actual = datetime.now().strftime("%A, %d de %B de %Y")

    context = {
        'total_consultas': total_consultas,
        'total_pacientes': total_pacientes,
        'consultas_hoy': consultas_hoy,
        'consultas_semana': consultas_semana,
        'consultas_mes': consultas_mes,
        'consultas_emergencia': consultas_emergencia,
        'duracion_promedio': round(duracion_promedio, 1),
        'usuario': usuario,
        'fecha_actual': fecha_actual,
        'ultimas_consultas': ultimas_consultas,
        'tipos_consulta': tipos_consulta,
        'pacientes_frecuentes': pacientes_frecuentes,
        'alertas': alertas[:10],  # Máximo 10 alertas
        'proximas_citas': proximas_citas,
    }
    return render(request, 'consultas/dashboard.html', context)


# ----------------------------
# Vista para listar todas las consultas
# ----------------------------
@login_required
def listar_consultas(request):
    # Filtros de búsqueda
    paciente_query = request.GET.get('paciente', '').strip()
    fecha_query = request.GET.get('fecha', '').strip()
    doctor_query = request.GET.get('doctor', '').strip()
    tipo_query = request.GET.get('tipo', '').strip()
    estado_query = request.GET.get('estado', '').strip()

    consultas = Consulta.objects.select_related('paciente', 'doctor').all().order_by('-fecha')

    # Aplicar filtros
    if paciente_query:
        consultas = consultas.filter(
            Q(paciente__primer_nombre__icontains=paciente_query) |
            Q(paciente__primer_apellido__icontains=paciente_query) |
            Q(paciente__documento_identificacion__icontains=paciente_query)
        )
    if fecha_query:
        consultas = consultas.filter(fecha__date=fecha_query)
    if doctor_query:
        consultas = consultas.filter(doctor__id=doctor_query)
    if tipo_query:
        consultas = consultas.filter(tipo_consulta=tipo_query)
    if estado_query:
        consultas = consultas.filter(estado=estado_query)

    # Paginación
    paginator = Paginator(consultas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Datos para filtros
    User = get_user_model()
    doctores = User.objects.all()
    tipos_consulta = Consulta.TIPO_CHOICES
    estados_consulta = Consulta.ESTADO_CHOICES

    # Estadísticas de filtros
    total_filtrado = consultas.count()
    consultas_emergencia = consultas.filter(tipo_consulta='Emergencia').count()
    consultas_pendientes = consultas.filter(estado='Programada').count()

    # Crear formulario para el modal
    form = ConsultaForm()

    context = {
        'consultas': page_obj,
        'paciente_query': paciente_query,
        'fecha_query': fecha_query,
        'doctor_query': doctor_query,
        'tipo_query': tipo_query,
        'estado_query': estado_query,
        'doctores': doctores,
        'tipos_consulta': tipos_consulta,
        'estados_consulta': estados_consulta,
        'page_obj': page_obj,
        'total_filtrado': total_filtrado,
        'consultas_emergencia': consultas_emergencia,
        'consultas_pendientes': consultas_pendientes,
        'form': form,  # Agregar formulario para el modal
    }
    return render(request, 'consultas/listar_consultas.html', context)


# ----------------------------
# Vista para editar una consulta
# ----------------------------
@login_required
def editar_consulta(request, consulta_id):
    consulta = get_object_or_404(Consulta, id=consulta_id)
    if consulta.doctor != request.user:
        messages.error(request, "🚫 No tienes permiso para editar esta consulta.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ConsultaForm(request.POST, instance=consulta)
        if form.is_valid():
            consulta = form.save(commit=False)
            consulta.fecha_actualizacion = timezone.now()
            
            # Calcular duración automáticamente si no se especifica
            if not consulta.duracion_consulta:
                consulta.duracion_consulta = consulta.calcular_duracion_consulta()
            
            consulta.save()
            messages.success(request, f'✏️ Consulta actualizada correctamente para {consulta.paciente.nombre_completo()}.')
            return redirect('listar_consultas')
        else:
            messages.error(request, "❌ Error al actualizar la consulta.")
    else:
        form = ConsultaForm(instance=consulta)
    
    # Obtener historial del paciente
    historial_consultas = consulta.obtener_historial_consultas()
    
    context = {
        'form': form,
        'consulta': consulta,
        'historial_consultas': historial_consultas,
    }
    return render(request, 'consultas/editar_consulta.html', context)


# ----------------------------
# Vista para eliminar una consulta
# ----------------------------
@login_required
def eliminar_consulta(request, consulta_id):
    consulta = get_object_or_404(Consulta, id=consulta_id)
    if consulta.doctor != request.user:
        messages.error(request, "🚫 No tienes permiso para eliminar esta consulta.")
        return redirect('dashboard')

    if request.method == 'POST':
        paciente_nombre = consulta.paciente.nombre_completo()
        consulta.delete()
        messages.success(request, f'🗑️ Consulta eliminada correctamente para {paciente_nombre}.')
        return redirect('listar_consultas')
    return render(request, 'consultas/eliminar_confirmar.html', {'consulta': consulta})


# ----------------------------
# Vista para ver detalles individuales de una consulta
# ----------------------------
@login_required
def detalle_consulta(request, consulta_id):
    consulta = get_object_or_404(Consulta, id=consulta_id)
    if consulta.doctor != request.user:
        messages.error(request, "🚫 No tienes permiso para ver esta consulta.")
        return redirect('dashboard')

    # Obtener datos adicionales
    historial_consultas = consulta.obtener_historial_consultas()
    estadisticas_vitales = consulta.obtener_estadisticas_vitales()
    alertas_medicas = consulta.obtener_alertas_medicas()

    context = {
        'consulta': consulta,
        'historial_consultas': historial_consultas,
        'estadisticas_vitales': estadisticas_vitales,
        'alertas_medicas': alertas_medicas,
    }
    return render(request, 'consultas/detalle_consulta.html', context)


# ----------------------------
# Vista para búsqueda rápida de pacientes (AJAX)
# ----------------------------
@login_required
def buscar_pacientes(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'pacientes': []})
    
    pacientes = Paciente.objects.filter(
        Q(primer_nombre__icontains=query) |
        Q(primer_apellido__icontains=query) |
        Q(documento_identificacion__icontains=query),
        estado='Activo'
    )[:10]
    
    resultados = []
    for paciente in pacientes:
        resultados.append({
            'id': paciente.id,
            'nombre': paciente.nombre_completo(),
            'documento': paciente.documento_identificacion,
            'edad': paciente.edad(),
            'telefono': paciente.telefono,
        })
    
    return JsonResponse({'pacientes': resultados})


# ----------------------------
# Vista para estadísticas de consultas
# ----------------------------
@login_required
def estadisticas_consultas(request):
    # Usar caché para mejorar rendimiento
    from django.core.cache import cache
    cache_key = f"estadisticas_{request.user.id}_{request.GET.urlencode()}"
    cached_result = cache.get(cache_key)
    
    if cached_result:
        return render(request, 'consultas/estadisticas.html', cached_result)
    # Filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Query base con límite para optimización
    consultas_query = Consulta.objects.all()
    
    # Si hay muchas consultas, limitar a los últimos 1000 registros para rendimiento
    total_consultas_sin_limite = consultas_query.count()
    if total_consultas_sin_limite > 1000:
        consultas_query = consultas_query.order_by('-fecha')[:1000]
        # Mostrar advertencia en el contexto
        limit_warning = f"Mostrando estadísticas de los últimos 1000 registros (de {total_consultas_sin_limite} total)"
    else:
        limit_warning = None
    
    # Aplicar filtros de fecha si se proporcionan
    if fecha_inicio:
        consultas_query = consultas_query.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        consultas_query = consultas_query.filter(fecha__lte=fecha_fin)
    
    # Estadísticas generales
    total_consultas = consultas_query.count()
    consultas_mes = consultas_query.filter(
        fecha__month=timezone.now().month
    ).count()
    
    # Promedios
    duracion_promedio = consultas_query.aggregate(
        promedio=Avg('duracion_consulta')
    )['promedio'] or 0
    
    costo_promedio = consultas_query.aggregate(
        promedio=Avg('costo_consulta')
    )['promedio'] or 0
    
    # Consultas por tipo
    consultas_por_tipo = consultas_query.values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Consultas por estado
    consultas_por_estado = consultas_query.values('estado').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Top doctores
    doctores_top = consultas_query.values('doctor__first_name', 'doctor__last_name').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Estadísticas adicionales
    consultas_hoy = consultas_query.filter(fecha__date=timezone.now().date()).count()
    consultas_semana = consultas_query.filter(
        fecha__gte=timezone.now().date() - timezone.timedelta(days=7)
    ).count()
    
    # Ingresos totales
    ingresos_totales = consultas_query.aggregate(
        total=Sum('costo_consulta')
    )['total'] or 0
    
    # Datos para gráficas (versión optimizada)
    from datetime import datetime, timedelta
    import calendar
    
    # Consultas por mes (últimos 6 meses) - optimizado
    datos_grafica_mensual = []
    meses_nombres = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    for i in range(5, -1, -1):  # Últimos 6 meses
        fecha_objetivo = timezone.now() - timedelta(days=30*i)
        mes_consultas = consultas_query.filter(
            fecha__year=fecha_objetivo.year,
            fecha__month=fecha_objetivo.month
        ).count()
        datos_grafica_mensual.append({
            'mes': f"{meses_nombres[fecha_objetivo.month-1]} {fecha_objetivo.year}",
            'total': mes_consultas
        })
    
    # Consultas por día de la semana - optimizado con SQLite
    datos_grafica_dias = []
    dias_nombres = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    
    # Usar consulta SQL directa para mejor rendimiento
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT strftime('%w', fecha) as dia_semana, COUNT(*) as total
            FROM consultas_consulta
            WHERE fecha IS NOT NULL
            GROUP BY strftime('%w', fecha)
            ORDER BY dia_semana
        """)
        resultados_dias = cursor.fetchall()
    
    # Crear diccionario con todos los días
    dias_count = {str(i): 0 for i in range(7)}
    for dia, total in resultados_dias:
        dias_count[dia] = total
    
    # Crear lista ordenada
    for i in range(7):
        datos_grafica_dias.append({
            'dia': dias_nombres[i],
            'total': dias_count[str(i)]
        })
    
    # Datos para gráfica de tipos de consulta (optimizado)
    datos_grafica_tipos = [
        {'tipo': tipo['tipo_consulta'], 'total': tipo['total']}
        for tipo in consultas_por_tipo[:10]  # Limitar a top 10
    ]
    
    # Datos para gráfica de estados (optimizado)
    datos_grafica_estados = [
        {'estado': estado['estado'], 'total': estado['total']}
        for estado in consultas_por_estado
    ]
    
    context = {
        'total_consultas': total_consultas,
        'consultas_mes': consultas_mes,
        'consultas_hoy': consultas_hoy,
        'consultas_semana': consultas_semana,
        'duracion_promedio': round(duracion_promedio, 1),
        'costo_promedio': round(costo_promedio, 2),
        'ingresos_totales': round(ingresos_totales, 2),
        'consultas_por_tipo': consultas_por_tipo,
        'consultas_por_estado': consultas_por_estado,
        'doctores_top': doctores_top,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'limit_warning': limit_warning,
        # Datos para gráficas
        'datos_grafica_mensual': datos_grafica_mensual,
        'datos_grafica_dias': datos_grafica_dias,
        'datos_grafica_tipos': datos_grafica_tipos,
        'datos_grafica_estados': datos_grafica_estados,
    }
    
    # Guardar en caché por 5 minutos
    cache.set(cache_key, context, 300)
    
    return render(request, 'consultas/estadisticas.html', context)


# ----------------------------
# Exportar estadísticas a Excel
# ----------------------------
@login_required
def exportar_estadisticas_excel(request):
    """Exporta las estadísticas de consultas a Excel"""
    try:
        from datetime import datetime
        
        # Aplicar filtros si se proporcionan
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        consultas_query = Consulta.objects.all()
        
        if fecha_inicio:
            consultas_query = consultas_query.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            consultas_query = consultas_query.filter(fecha__lte=fecha_fin)
        
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen General
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Estadísticas generales
        total_consultas = consultas_query.count()
        consultas_mes = consultas_query.filter(fecha__month=timezone.now().month).count()
        duracion_promedio = consultas_query.aggregate(promedio=Avg('duracion_consulta'))['promedio'] or 0
        costo_promedio = consultas_query.aggregate(promedio=Avg('costo_consulta'))['promedio'] or 0
        
        # Título principal
        ws_resumen.merge_cells('A1:D1')
        ws_resumen['A1'] = 'REPORTE ESTADÍSTICO DEL SISTEMA MÉDICO'
        ws_resumen['A1'].font = openpyxl.styles.Font(size=16, bold=True, color="1E3C72")
        ws_resumen['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        ws_resumen.merge_cells('A2:D2')
        ws_resumen['A2'] = 'Análisis Integral de Consultas Médicas - HealthLife'
        ws_resumen['A2'].font = openpyxl.styles.Font(size=12, color="666666")
        ws_resumen['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        ws_resumen.append([])
        
        # Información de filtros aplicados
        if fecha_inicio or fecha_fin:
            ws_resumen.merge_cells('A4:D4')
            ws_resumen['A4'] = 'FILTROS APLICADOS'
            ws_resumen['A4'].font = openpyxl.styles.Font(size=12, bold=True, color="1E3C72")
            ws_resumen['A4'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            if fecha_inicio:
                ws_resumen.append(['Fecha de Inicio', fecha_inicio, '', ''])
            if fecha_fin:
                ws_resumen.append(['Fecha de Fin', fecha_fin, '', ''])
            ws_resumen.append([])
        
        # Estadísticas generales
        ws_resumen.merge_cells(f'A{ws_resumen.max_row + 1}:D{ws_resumen.max_row + 1}')
        ws_resumen[f'A{ws_resumen.max_row}'] = 'ESTADÍSTICAS GENERALES'
        ws_resumen[f'A{ws_resumen.max_row}'].font = openpyxl.styles.Font(size=14, bold=True, color="1E3C72")
        ws_resumen[f'A{ws_resumen.max_row}'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        ws_resumen.append([])
        ws_resumen.append(['Métrica', 'Valor', 'Descripción', ''])
        
        # Calcular estadísticas adicionales para el reporte
        consultas_hoy = consultas_query.filter(fecha__date=timezone.now().date()).count()
        consultas_semana = consultas_query.filter(
            fecha__gte=timezone.now().date() - timezone.timedelta(days=7)
        ).count()
        ingresos_totales = consultas_query.aggregate(total=Sum('costo_consulta'))['total'] or 0
        doctores_activos = consultas_query.values('doctor').distinct().count()
        
        # Datos para gráficas
        from datetime import datetime, timedelta
        import calendar
        
        # Consultas por mes (últimos 6 meses)
        datos_grafica_mensual = []
        meses_nombres = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        
        for i in range(5, -1, -1):
            fecha_objetivo = timezone.now() - timedelta(days=30*i)
            mes_consultas = consultas_query.filter(
                fecha__year=fecha_objetivo.year,
                fecha__month=fecha_objetivo.month
            ).count()
            datos_grafica_mensual.append({
                'mes': f"{meses_nombres[fecha_objetivo.month-1]} {fecha_objetivo.year}",
                'total': mes_consultas
            })
        
        # Consultas por día de la semana
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT strftime('%w', fecha) as dia_semana, COUNT(*) as total
                FROM consultas_consulta
                WHERE fecha IS NOT NULL
                GROUP BY strftime('%w', fecha)
                ORDER BY dia_semana
            """)
            resultados_dias = cursor.fetchall()
        
        dias_nombres = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        datos_grafica_dias = []
        dias_count = {str(i): 0 for i in range(7)}
        for dia, total in resultados_dias:
            dias_count[dia] = total
        
        for i in range(7):
            datos_grafica_dias.append({
                'dia': dias_nombres[i],
                'total': dias_count[str(i)]
            })
        
        # Datos con formato
        datos_estadisticas = [
            ['Total de Consultas', total_consultas, 'Número total de consultas registradas', ''],
            ['Consultas del Mes', consultas_mes, 'Consultas realizadas en el mes actual', ''],
            ['Consultas de Hoy', consultas_hoy, 'Consultas realizadas hoy', ''],
            ['Consultas de la Semana', consultas_semana, 'Consultas de los últimos 7 días', ''],
            ['Duración Promedio (min)', round(duracion_promedio, 1), 'Tiempo promedio por consulta', ''],
            ['Costo Promedio (Q)', round(costo_promedio, 2), 'Costo promedio por consulta', ''],
            ['Ingresos Totales (Q)', round(ingresos_totales, 2), 'Ingresos totales generados', ''],
            ['Doctores Activos', doctores_activos, 'Número de doctores con actividad', ''],
            ['Fecha de Reporte', datetime.now().strftime('%d/%m/%Y %H:%M'), 'Fecha y hora de generación', ''],
        ]
        
        for dato in datos_estadisticas:
            ws_resumen.append(dato)
        
        # Aplicar formato a la tabla
        for row in ws_resumen.iter_rows(min_row=ws_resumen.max_row-len(datos_estadisticas), max_row=ws_resumen.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        # Hoja 2: Gráficas de Barras - Consultas por Mes
        ws_grafica_mensual = wb.create_sheet("Gráfica Mensual")
        ws_grafica_mensual['A1'] = 'GRÁFICA DE BARRAS - CONSULTAS POR MES'
        ws_grafica_mensual['A1'].font = openpyxl.styles.Font(size=14, bold=True, color="1E3C72")
        ws_grafica_mensual.append([])
        ws_grafica_mensual.append(['Mes', 'Total de Consultas'])
        
        for dato in datos_grafica_mensual:
            ws_grafica_mensual.append([dato['mes'], dato['total']])
        
        # Crear gráfica de barras real
        from openpyxl.chart import BarChart, Reference
        
        chart_mensual = BarChart()
        chart_mensual.title = "Consultas por Mes"
        chart_mensual.style = 10
        chart_mensual.x_axis.title = "Mes"
        chart_mensual.y_axis.title = "Total de Consultas"
        
        data = Reference(ws_grafica_mensual, min_col=2, min_row=3, max_row=ws_grafica_mensual.max_row)
        cats = Reference(ws_grafica_mensual, min_col=1, min_row=3, max_row=ws_grafica_mensual.max_row)
        chart_mensual.add_data(data, titles_from_data=True)
        chart_mensual.set_categories(cats)
        
        # Agregar gráfica a la hoja
        ws_grafica_mensual.add_chart(chart_mensual, "D2")
        
        # Aplicar formato a la tabla
        for row in ws_grafica_mensual.iter_rows(min_row=3, max_row=ws_grafica_mensual.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        # Hoja 3: Gráficas de Barras - Consultas por Día
        ws_grafica_dias = wb.create_sheet("Gráfica Días")
        ws_grafica_dias['A1'] = 'GRÁFICA DE BARRAS - CONSULTAS POR DÍA DE LA SEMANA'
        ws_grafica_dias['A1'].font = openpyxl.styles.Font(size=14, bold=True, color="1E3C72")
        ws_grafica_dias.append([])
        ws_grafica_dias.append(['Día', 'Total de Consultas'])
        
        for dato in datos_grafica_dias:
            ws_grafica_dias.append([dato['dia'], dato['total']])
        
        # Crear gráfica de barras real para días
        chart_dias = BarChart()
        chart_dias.title = "Consultas por Día de la Semana"
        chart_dias.style = 11
        chart_dias.x_axis.title = "Día"
        chart_dias.y_axis.title = "Total de Consultas"
        
        data_dias = Reference(ws_grafica_dias, min_col=2, min_row=3, max_row=ws_grafica_dias.max_row)
        cats_dias = Reference(ws_grafica_dias, min_col=1, min_row=3, max_row=ws_grafica_dias.max_row)
        chart_dias.add_data(data_dias, titles_from_data=True)
        chart_dias.set_categories(cats_dias)
        
        # Agregar gráfica a la hoja
        ws_grafica_dias.add_chart(chart_dias, "D2")
        
        # Aplicar formato a la tabla
        for row in ws_grafica_dias.iter_rows(min_row=3, max_row=ws_grafica_dias.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        # Hoja 4: Consultas por Tipo
        ws_tipo = wb.create_sheet("Consultas por Tipo")
        
        # Título
        ws_tipo.merge_cells('A1:C1')
        ws_tipo['A1'] = 'DISTRIBUCIÓN POR TIPO DE CONSULTA'
        ws_tipo['A1'].font = openpyxl.styles.Font(size=14, bold=True, color="1E3C72")
        ws_tipo['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        ws_tipo.append([])
        ws_tipo.append(['Tipo de Consulta', 'Cantidad', 'Porcentaje'])
        
        # Encabezados con formato
        for col in ['A', 'B', 'C']:
            ws_tipo[f'{col}{ws_tipo.max_row}'].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            ws_tipo[f'{col}{ws_tipo.max_row}'].fill = openpyxl.styles.PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
        
        consultas_por_tipo = consultas_query.values('tipo_consulta').annotate(
            total=Count('id')
        ).order_by('-total')
        
        for tipo in consultas_por_tipo:
            porcentaje = (tipo['total'] / total_consultas * 100) if total_consultas > 0 else 0
            ws_tipo.append([tipo['tipo_consulta'], tipo['total'], f"{porcentaje:.1f}%"])
        
        # Crear gráfica de pastel para tipos de consulta
        from openpyxl.chart import PieChart
        
        chart_tipos = PieChart()
        chart_tipos.title = "Distribución por Tipo de Consulta"
        
        data_tipos = Reference(ws_tipo, min_col=2, min_row=3, max_row=ws_tipo.max_row)
        cats_tipos = Reference(ws_tipo, min_col=1, min_row=3, max_row=ws_tipo.max_row)
        chart_tipos.add_data(data_tipos, titles_from_data=True)
        chart_tipos.set_categories(cats_tipos)
        
        # Agregar gráfica a la hoja
        ws_tipo.add_chart(chart_tipos, "E2")
        
        # Hoja 5: Consultas por Estado
        ws_estado = wb.create_sheet("Consultas por Estado")
        ws_estado.append(['Estado', 'Cantidad', 'Porcentaje'])
        
        consultas_por_estado = consultas_query.values('estado').annotate(
            total=Count('id')
        ).order_by('-total')
        
        for estado in consultas_por_estado:
            porcentaje = (estado['total'] / total_consultas * 100) if total_consultas > 0 else 0
            ws_estado.append([estado['estado'], estado['total'], f"{porcentaje:.1f}%"])
        
        # Crear gráfica de pastel para estados
        chart_estados = PieChart()
        chart_estados.title = "Distribución por Estado de Consulta"
        
        data_estados = Reference(ws_estado, min_col=2, min_row=3, max_row=ws_estado.max_row)
        cats_estados = Reference(ws_estado, min_col=1, min_row=3, max_row=ws_estado.max_row)
        chart_estados.add_data(data_estados, titles_from_data=True)
        chart_estados.set_categories(cats_estados)
        
        # Agregar gráfica a la hoja
        ws_estado.add_chart(chart_estados, "E2")
        
        # Hoja 6: Top Doctores
        ws_doctores = wb.create_sheet("Top Doctores")
        ws_doctores.append(['Doctor', 'Consultas', 'Porcentaje'])
        
        doctores_top = consultas_query.values('doctor__first_name', 'doctor__last_name').annotate(
            total=Count('id')
        ).order_by('-total')[:10]
        
        for doctor in doctores_top:
            nombre = f"{doctor['doctor__first_name'] or ''} {doctor['doctor__last_name'] or ''}".strip()
            porcentaje = (doctor['total'] / total_consultas * 100) if total_consultas > 0 else 0
            ws_doctores.append([nombre, doctor['total'], f"{porcentaje:.1f}%"])
        
        # Crear gráfica de barras para doctores
        chart_doctores = BarChart()
        chart_doctores.title = "Rendimiento de Doctores"
        chart_doctores.style = 12
        chart_doctores.x_axis.title = "Doctor"
        chart_doctores.y_axis.title = "Total de Consultas"
        
        data_doctores = Reference(ws_doctores, min_col=2, min_row=3, max_row=ws_doctores.max_row)
        cats_doctores = Reference(ws_doctores, min_col=1, min_row=3, max_row=ws_doctores.max_row)
        chart_doctores.add_data(data_doctores, titles_from_data=True)
        chart_doctores.set_categories(cats_doctores)
        
        # Agregar gráfica a la hoja
        ws_doctores.add_chart(chart_doctores, "E2")
        
        # Hoja 7: Consultas por Mes (últimos 12 meses)
        ws_mes = wb.create_sheet("Consultas por Mes")
        ws_mes.append(['Mes', 'Año', 'Consultas'])
        
        from django.db.models.functions import TruncMonth
        consultas_mensuales = consultas_query.annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Count('id')
        ).order_by('-mes')[:12]
        
        for mes in consultas_mensuales:
            ws_mes.append([
                mes['mes'].strftime('%B'),
                mes['mes'].year,
                mes['total']
            ])
        
        # Crear gráfica de línea para consultas por mes
        from openpyxl.chart import LineChart
        
        chart_mes = LineChart()
        chart_mes.title = "Tendencia de Consultas por Mes"
        chart_mes.style = 13
        chart_mes.x_axis.title = "Mes"
        chart_mes.y_axis.title = "Total de Consultas"
        
        data_mes = Reference(ws_mes, min_col=3, min_row=3, max_row=ws_mes.max_row)
        cats_mes = Reference(ws_mes, min_col=1, min_row=3, max_row=ws_mes.max_row)
        chart_mes.add_data(data_mes, titles_from_data=True)
        chart_mes.set_categories(cats_mes)
        
        # Agregar gráfica a la hoja
        ws_mes.add_chart(chart_mes, "E2")
        
        # Configurar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="estadisticas_consultas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Error al exportar estadísticas: {str(e)}")
        return redirect('estadisticas_consultas')


# ----------------------------
# Exportar estadísticas a PDF profesional
# ----------------------------
@login_required
def exportar_estadisticas_pdf(request):
    """Exporta las estadísticas de consultas a PDF con formato profesional"""
    from datetime import datetime
    
    # Aplicar filtros si se proporcionan
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    consultas_query = Consulta.objects.all()
    
    if fecha_inicio:
        consultas_query = consultas_query.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        consultas_query = consultas_query.filter(fecha__lte=fecha_fin)
    
    # Estadísticas generales
    total_consultas = consultas_query.count()
    consultas_mes = consultas_query.filter(fecha__month=timezone.now().month).count()
    duracion_promedio = consultas_query.aggregate(promedio=Avg('duracion_consulta'))['promedio'] or 0
    costo_promedio = consultas_query.aggregate(promedio=Avg('costo_consulta'))['promedio'] or 0
    
    # Estadísticas adicionales
    consultas_hoy = consultas_query.filter(fecha__date=timezone.now().date()).count()
    consultas_semana = consultas_query.filter(
        fecha__gte=timezone.now().date() - timezone.timedelta(days=7)
    ).count()
    
    # Ingresos totales
    ingresos_totales = consultas_query.aggregate(
        total=Sum('costo_consulta')
    )['total'] or 0
    
    # Consultas por tipo
    consultas_por_tipo = consultas_query.values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Consultas por estado
    consultas_por_estado = consultas_query.values('estado').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Top doctores
    doctores_top = consultas_query.values('doctor__first_name', 'doctor__last_name').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    # Calcular datos adicionales para gráficas
    from datetime import datetime, timedelta
    import calendar
    
    # Consultas por mes (últimos 6 meses)
    datos_grafica_mensual = []
    meses_nombres = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    for i in range(5, -1, -1):
        fecha_objetivo = timezone.now() - timedelta(days=30*i)
        mes_consultas = consultas_query.filter(
            fecha__year=fecha_objetivo.year,
            fecha__month=fecha_objetivo.month
        ).count()
        datos_grafica_mensual.append({
            'mes': f"{meses_nombres[fecha_objetivo.month-1]} {fecha_objetivo.year}",
            'total': mes_consultas
        })
    
    # Consultas por día de la semana
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT strftime('%w', fecha) as dia_semana, COUNT(*) as total
            FROM consultas_consulta
            WHERE fecha IS NOT NULL
            GROUP BY strftime('%w', fecha)
            ORDER BY dia_semana
        """)
        resultados_dias = cursor.fetchall()
    
    dias_nombres = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    datos_grafica_dias = []
    dias_count = {str(i): 0 for i in range(7)}
    for dia, total in resultados_dias:
        dias_count[dia] = total
    
    for i in range(7):
        datos_grafica_dias.append({
            'dia': dias_nombres[i],
            'total': dias_count[str(i)]
        })
    
    # Calcular valores máximos para las gráficas
    max_value = max([d['total'] for d in datos_grafica_mensual]) if datos_grafica_mensual else 1
    max_value_dias = max([d['total'] for d in datos_grafica_dias]) if datos_grafica_dias else 1
    
    # Datos adicionales para análisis
    distribucion_tendencia = "estable" if max_value < 50 else "creciente"
    dia_mas_ocupado = max(datos_grafica_dias, key=lambda x: x['total'])['dia'] if datos_grafica_dias else "N/A"
    max_consultas_dia = max([d['total'] for d in datos_grafica_dias]) if datos_grafica_dias else 0
    tipo_mas_comun = consultas_por_tipo[0]['tipo_consulta'] if consultas_por_tipo else "N/A"
    
    context = {
        'total_consultas': total_consultas,
        'consultas_mes': consultas_mes,
        'consultas_hoy': consultas_hoy,
        'consultas_semana': consultas_semana,
        'duracion_promedio': round(duracion_promedio, 1),
        'costo_promedio': round(costo_promedio, 2),
        'ingresos_totales': round(ingresos_totales, 2),
        'consultas_por_tipo': consultas_por_tipo,
        'consultas_por_estado': consultas_por_estado,
        'doctores_top': doctores_top,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'fecha_actual': datetime.now().strftime('%d/%m/%Y %H:%M'),
        # Datos para gráficas
        'datos_grafica_mensual': datos_grafica_mensual,
        'datos_grafica_dias': datos_grafica_dias,
        'max_value': max_value,
        'max_value_dias': max_value_dias,
        # Datos para análisis
        'distribucion_tendencia': distribucion_tendencia,
        'dia_mas_ocupado': dia_mas_ocupado,
        'max_consultas_dia': max_consultas_dia,
        'tipo_mas_comun': tipo_mas_comun,
    }
    
    template_path = 'consultas/pdf_estadisticas_simple.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_estadistico_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    pisa.CreatePDF(html, dest=response)
    return response


# ----------------------------
# Exportar consultas a PDF mejorado
# ----------------------------
@login_required
def exportar_consultas_pdf(request):
    from datetime import datetime
    consultas = Consulta.objects.select_related('paciente', 'doctor').all().order_by('-fecha')
    template_path = 'consultas/pdf_consultas.html'
    context = {
        'consultas': consultas,
        'fecha_actual': datetime.now().strftime('%d/%m/%Y %H:%M')
    }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="consultas_medicas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa.CreatePDF(html, dest=response)
    return response


# ----------------------------
# Exportar consultas a Excel mejorado
# ----------------------------
@login_required
def exportar_consultas_excel(request):
    from datetime import datetime
    consultas = Consulta.objects.select_related('paciente', 'doctor').all().order_by('-fecha')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consultas Médicas"

    # Encabezados mejorados
    ws.append([
        '#', 'Paciente', 'Documento', 'Edad', 'Motivo', 'Diagnóstico', 'Tratamiento',
        'Tipo', 'Estado', 'Signos Vitales', 'IMC', 'Fecha', 'Doctor', 'Duración (min)', 'Costo'
    ])

    for idx, c in enumerate(consultas, 1):
        # Preparar signos vitales
        signos_vitales = []
        if c.presion_arterial:
            signos_vitales.append(f"PA: {c.presion_arterial}")
        if c.temperatura:
            signos_vitales.append(f"Temp: {c.temperatura}°C")
        if c.frecuencia_cardiaca:
            signos_vitales.append(f"FC: {c.frecuencia_cardiaca} lpm")
        
        signos_str = " | ".join(signos_vitales) if signos_vitales else "No registrados"
        
        # Verificar que el paciente existe
        paciente_nombre = c.paciente.nombre_completo() if c.paciente else "Paciente no encontrado"
        paciente_documento = c.paciente.documento_identificacion if c.paciente else "N/A"
        paciente_edad = c.paciente.edad() if c.paciente else "N/A"
        
        ws.append([
            idx,
            paciente_nombre,
            paciente_documento,
            paciente_edad,
            c.motivo or "N/A",
            c.diagnostico or "N/A",
            c.tratamiento or "N/A",
            c.tipo_consulta or "N/A",
            c.estado or "N/A",
            signos_str,
            c.calcular_imc() or "N/A",
            c.fecha.strftime("%d/%m/%Y %H:%M") if c.fecha else "N/A",
            c.doctor.get_full_name() or c.doctor.username if c.doctor else "N/A",
            c.duracion_consulta or "N/A",
            c.costo_consulta or "0.00"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="consultas_medicas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


# ----------------------------
# Función de prueba para estadísticas
# ----------------------------
@login_required
def test_estadisticas(request):
    """Función de prueba para verificar que las estadísticas funcionen"""
    try:
        # Consulta simple
        total_consultas = Consulta.objects.count()
        
        context = {
            'total_consultas': total_consultas,
            'mensaje': 'Estadísticas funcionando correctamente',
            'fecha': timezone.now().strftime('%d/%m/%Y %H:%M')
        }
        
        return render(request, 'consultas/test_estadisticas.html', context)
        
    except Exception as e:
        return HttpResponse(f"Error en estadísticas: {str(e)}", content_type='text/plain')


# ----------------------------
# Función de prueba para exportaciones
# ----------------------------
@login_required
def test_export(request):
    """Función de prueba para verificar que las exportaciones funcionen"""
    try:
        # Probar exportación de estadísticas
        from datetime import datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prueba"
        ws.append(['Test', 'Funcionando', datetime.now().strftime('%d/%m/%Y %H:%M')])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_export.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error en exportación: {str(e)}", content_type='text/plain')


@login_required
def test_export_page(request):
    """Página de prueba para verificar exportaciones"""
    try:
        total_consultas = Consulta.objects.count()
        
        context = {
            'total_consultas': total_consultas,
            'mensaje': 'Sistema funcionando correctamente',
            'fecha': timezone.now().strftime('%d/%m/%Y %H:%M')
        }
        
        return render(request, 'consultas/test_export.html', context)
        
    except Exception as e:
        context = {
            'total_consultas': 0,
            'mensaje': f'Error en el sistema: {str(e)}',
            'fecha': timezone.now().strftime('%d/%m/%Y %H:%M')
        }
        return render(request, 'consultas/test_export.html', context)


# ----------------------------
# Manejador personalizado de error 403
# ----------------------------
def error_403(request, exception=None):
    return render(request, '403.html', status=403)


# ----------------------------
# Vistas para WhatsApp Notifications
# ----------------------------
@login_required
def enviar_receta_whatsapp(request, consulta_id):
    """Vista para enviar receta médica por WhatsApp"""
    consulta = get_object_or_404(Consulta, id=consulta_id)
    
    if request.method == 'POST':
        try:
            from .services import WhatsAppNotificationService
            
            service = WhatsAppNotificationService()
            notificacion = service.send_prescription(consulta, enviado_por=request.user)
            
            messages.success(request, f"✅ Receta enviada exitosamente por WhatsApp a {consulta.paciente.nombre_completo()}")
            return JsonResponse({
                'success': True,
                'message': 'Receta enviada exitosamente',
                'notification_id': notificacion.id
            })
            
        except Exception as e:
            messages.error(request, f"❌ Error al enviar receta por WhatsApp: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    # Verificar si el paciente tiene habilitadas las notificaciones
    if not consulta.paciente.notificar_whatsapp:
        messages.warning(request, f"⚠️ El paciente {consulta.paciente.nombre_completo()} no tiene habilitadas las notificaciones por WhatsApp")
    
    context = {
        'consulta': consulta,
        'paciente': consulta.paciente,
        'whatsapp_enabled': consulta.paciente.notificar_whatsapp
    }
    return render(request, 'consultas/enviar_receta_whatsapp.html', context)


@login_required
def enviar_resumen_whatsapp(request, consulta_id):
    """Vista para enviar resumen de consulta por WhatsApp"""
    consulta = get_object_or_404(Consulta, id=consulta_id)
    
    if request.method == 'POST':
        try:
            from .services import WhatsAppNotificationService
            
            service = WhatsAppNotificationService()
            notificacion = service.send_consultation_summary(consulta, enviado_por=request.user)
            
            messages.success(request, f"✅ Resumen de consulta enviado exitosamente por WhatsApp a {consulta.paciente.nombre_completo()}")
            return JsonResponse({
                'success': True,
                'message': 'Resumen enviado exitosamente',
                'notification_id': notificacion.id
            })
            
        except Exception as e:
            messages.error(request, f"❌ Error al enviar resumen por WhatsApp: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    # Verificar si el paciente tiene habilitadas las notificaciones
    if not consulta.paciente.notificar_whatsapp:
        messages.warning(request, f"⚠️ El paciente {consulta.paciente.nombre_completo()} no tiene habilitadas las notificaciones por WhatsApp")
    
    context = {
        'consulta': consulta,
        'paciente': consulta.paciente,
        'whatsapp_enabled': consulta.paciente.notificar_whatsapp
    }
    return render(request, 'consultas/enviar_resumen_whatsapp.html', context)


@login_required
def listar_notificaciones_whatsapp(request):
    """Vista para listar todas las notificaciones de WhatsApp"""
    # Filtros de búsqueda
    paciente_query = request.GET.get('paciente', '').strip()
    tipo_query = request.GET.get('tipo', '').strip()
    estado_query = request.GET.get('estado', '').strip()
    fecha_query = request.GET.get('fecha', '').strip()
    
    notificaciones = NotificacionWhatsApp.objects.select_related(
        'paciente', 'consulta', 'enviado_por'
    ).all().order_by('-fecha_envio')
    
    # Aplicar filtros
    if paciente_query:
        notificaciones = notificaciones.filter(
            Q(paciente__primer_nombre__icontains=paciente_query) |
            Q(paciente__primer_apellido__icontains=paciente_query) |
            Q(paciente__documento_identificacion__icontains=paciente_query)
        )
    if tipo_query:
        notificaciones = notificaciones.filter(tipo=tipo_query)
    if estado_query:
        notificaciones = notificaciones.filter(estado=estado_query)
    if fecha_query:
        notificaciones = notificaciones.filter(fecha_envio__date=fecha_query)
    
    # Paginación
    paginator = Paginator(notificaciones, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_notificaciones = notificaciones.count()
    notificaciones_exitosas = notificaciones.filter(estado__in=['entregado', 'leido']).count()
    notificaciones_pendientes = notificaciones.filter(estado='pendiente').count()
    notificaciones_error = notificaciones.filter(estado='error').count()
    
    # Tipos y estados para filtros
    tipos_notificacion = NotificacionWhatsApp.TIPO_CHOICES
    estados_notificacion = NotificacionWhatsApp.ESTADO_CHOICES
    
    context = {
        'page_obj': page_obj,
        'total_notificaciones': total_notificaciones,
        'notificaciones_exitosas': notificaciones_exitosas,
        'notificaciones_pendientes': notificaciones_pendientes,
        'notificaciones_error': notificaciones_error,
        'tipos_notificacion': tipos_notificacion,
        'estados_notificacion': estados_notificacion,
        'filtros': {
            'paciente': paciente_query,
            'tipo': tipo_query,
            'estado': estado_query,
            'fecha': fecha_query,
        }
    }
    
    return render(request, 'consultas/listar_notificaciones_whatsapp.html', context)


@login_required
def detalle_notificacion_whatsapp(request, notificacion_id):
    """Vista para ver el detalle de una notificación de WhatsApp"""
    notificacion = get_object_or_404(NotificacionWhatsApp, id=notificacion_id)
    
    context = {
        'notificacion': notificacion,
        'estadisticas': notificacion.obtener_estadisticas(),
    }
    
    return render(request, 'consultas/detalle_notificacion_whatsapp.html', context)


@login_required
def reintentar_notificacion_whatsapp(request, notificacion_id):
    """Vista para reintentar el envío de una notificación fallida"""
    notificacion = get_object_or_404(NotificacionWhatsApp, id=notificacion_id)
    
    if request.method == 'POST':
        if not notificacion.puede_reintentar():
            messages.error(request, "❌ No se puede reintentar esta notificación")
            return JsonResponse({
                'success': False,
                'message': 'No se puede reintentar esta notificación'
            }, status=400)
        
        try:
            from .services import WhatsAppNotificationService
            
            service = WhatsAppNotificationService()
            
            if notificacion.tipo == 'receta' and notificacion.consulta:
                nueva_notificacion = service.send_prescription(notificacion.consulta, enviado_por=request.user)
            elif notificacion.tipo == 'consulta' and notificacion.consulta:
                nueva_notificacion = service.send_consultation_summary(notificacion.consulta, enviado_por=request.user)
            else:
                raise Exception("Tipo de notificación no soportado para reintento")
            
            # Marcar la notificación original como cancelada
            notificacion.estado = 'cancelado'
            notificacion.save()
            
            messages.success(request, f"✅ Notificación reenviada exitosamente")
            return JsonResponse({
                'success': True,
                'message': 'Notificación reenviada exitosamente',
                'notification_id': nueva_notificacion.id
            })
            
        except Exception as e:
            messages.error(request, f"❌ Error al reenviar notificación: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    context = {
        'notificacion': notificacion,
    }
    return render(request, 'consultas/reintentar_notificacion_whatsapp.html', context)


@login_required
def estadisticas_whatsapp(request):
    """Vista para mostrar estadísticas de notificaciones de WhatsApp"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Período de análisis (últimos 30 días por defecto)
    dias = int(request.GET.get('dias', 30))
    fecha_inicio = timezone.now() - timedelta(days=dias)
    
    # Notificaciones del período
    notificaciones_periodo = NotificacionWhatsApp.objects.filter(
        fecha_envio__gte=fecha_inicio
    )
    
    # Estadísticas generales
    total_notificaciones = notificaciones_periodo.count()
    notificaciones_exitosas = notificaciones_periodo.filter(
        estado__in=['entregado', 'leido']
    ).count()
    tasa_exito = (notificaciones_exitosas / total_notificaciones * 100) if total_notificaciones > 0 else 0
    
    # Estadísticas por tipo
    stats_por_tipo = notificaciones_periodo.values('tipo').annotate(
        total=Count('id'),
        exitosas=Count('id', filter=Q(estado__in=['entregado', 'leido'])),
        errores=Count('id', filter=Q(estado='error'))
    )
    
    # Estadísticas por día (últimos 7 días)
    stats_por_dia = []
    for i in range(7):
        fecha = timezone.now().date() - timedelta(days=i)
        count = notificaciones_periodo.filter(fecha_envio__date=fecha).count()
        stats_por_dia.append({
            'fecha': fecha.strftime('%d/%m'),
            'cantidad': count
        })
    stats_por_dia.reverse()
    
    # Top pacientes con más notificaciones
    top_pacientes = notificaciones_periodo.values(
        'paciente__primer_nombre', 'paciente__primer_apellido'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    # Top doctores que envían más notificaciones
    top_doctores = notificaciones_periodo.values(
        'enviado_por__first_name', 'enviado_por__last_name'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    context = {
        'dias': dias,
        'fecha_inicio': fecha_inicio,
        'total_notificaciones': total_notificaciones,
        'notificaciones_exitosas': notificaciones_exitosas,
        'tasa_exito': round(tasa_exito, 2),
        'stats_por_tipo': stats_por_tipo,
        'stats_por_dia': stats_por_dia,
        'top_pacientes': top_pacientes,
        'top_doctores': top_doctores,
    }
    
    return render(request, 'consultas/estadisticas_whatsapp.html', context)


@login_required
def verificar_configuracion_whatsapp(request):
    """Vista para verificar la configuración de WhatsApp"""
    from .services import WhatsAppService
    from decouple import config
    
    context = {
        'configuracion': {},
        'errores': [],
        'advertencias': [],
        'exitoso': False
    }
    
    try:
        # Verificar variables de entorno
        api_token = config('WHATSAPP_API_TOKEN', default='')
        phone_number_id = config('WHATSAPP_PHONE_NUMBER_ID', default='')
        business_account_id = config('WHATSAPP_BUSINESS_ACCOUNT_ID', default='')
        app_id = config('WHATSAPP_APP_ID', default='')
        enabled = config('WHATSAPP_ENABLED', default=False, cast=bool)
        
        context['configuracion'] = {
            'api_token': '✅ Configurado' if api_token else '❌ No configurado',
            'phone_number_id': '✅ Configurado' if phone_number_id else '❌ No configurado',
            'business_account_id': '✅ Configurado' if business_account_id else '❌ No configurado',
            'app_id': '✅ Configurado' if app_id else '❌ No configurado',
            'enabled': '✅ Habilitado' if enabled else '❌ Deshabilitado',
        }
        
        # Verificar si el servicio está habilitado
        whatsapp_service = WhatsAppService()
        
        if whatsapp_service.is_enabled():
            context['exitoso'] = True
            context['mensaje'] = '✅ WhatsApp está configurado correctamente'
        else:
            context['errores'].append('❌ WhatsApp no está habilitado. Verifica las credenciales.')
            
    except Exception as e:
        context['errores'].append(f'❌ Error al verificar configuración: {str(e)}')
    
    # Verificar si hay pacientes con WhatsApp habilitado
    from pacientes.models import Paciente
    pacientes_whatsapp = Paciente.objects.filter(notificar_whatsapp=True).count()
    
    if pacientes_whatsapp == 0:
        context['advertencias'].append('⚠️ No hay pacientes con WhatsApp habilitado')
    else:
        context['configuracion']['pacientes_whatsapp'] = f'✅ {pacientes_whatsapp} pacientes con WhatsApp habilitado'
    
    return render(request, 'consultas/verificar_configuracion_whatsapp.html', context)


@login_required
def prueba_whatsapp_demo(request):
    """Vista para probar WhatsApp en modo demo"""
    if request.method == 'POST':
        # Simular envío de mensaje
        from .services import WhatsAppService
        
        whatsapp_service = WhatsAppService()
        
        # Datos de prueba
        telefono_prueba = request.POST.get('telefono', '50212345678')
        mensaje_prueba = request.POST.get('mensaje', 'Este es un mensaje de prueba desde HealthLife')
        
        try:
            # Simular envío
            response = whatsapp_service.send_text_message(telefono_prueba, mensaje_prueba)
            
            messages.success(request, f'✅ Mensaje de prueba enviado (MODO DEMO)')
            messages.info(request, f'📱 Número: {telefono_prueba}')
            messages.info(request, f'💬 Mensaje: {mensaje_prueba}')
            messages.warning(request, '⚠️ Este es un mensaje simulado - No llegará realmente a WhatsApp')
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
    
    return render(request, 'consultas/prueba_whatsapp_demo.html')
